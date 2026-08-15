import json
import zipfile

from openpyxl import load_workbook

from qpcr_analyzer.analysis import analyze
from qpcr_analyzer.exports import export_all_svg, export_clean_excel, make_gene_figure
from qpcr_analyzer.models import AnalysisConfig, GeneDefinition, PlateAssignment, ProjectState, WellRecord
from qpcr_analyzer.project import load_project, save_project


def sample_state(tmp_path):
    source = tmp_path/'source.csv'; source.write_text('Pos,Cp', encoding='utf-8')
    wells, assignments = [], {}
    layout = [('A', 'Ctrl-1', 'Ctrl', 20.0), ('B', 'Ctrl-2', 'Ctrl', 20.2),
              ('C', 'Model-1', 'Model', 18.0), ('D', 'Model-2', 'Model', 18.2)]
    for plate_row, sample, group, target in layout:
        for index, ct in enumerate([target, target+.1, target-.1], 1):
            well = f'{plate_row}{index}'; wells.append(WellRecord(well, ct))
            assignments[well] = PlateAssignment(well, sample, group, 'Target', False, index, plate_row=plate_row)
        for index, ct in enumerate([17.0, 17.1, 16.9], 4):
            well = f'{plate_row}{index}'; wells.append(WellRecord(well, ct))
            assignments[well] = PlateAssignment(well, sample, group, 'REF', True, index-3, plate_row=plate_row)
    return ProjectState(source_path=str(source), source_sheet='Sheet1', header_row=1,
        wells=tuple(wells), assignments=assignments, config=AnalysisConfig(96),
        genes=[GeneDefinition('Target', False, 0), GeneDefinition('REF', True, 1)])


def test_project_v2_round_trip(tmp_path):
    state = sample_state(tmp_path); state.exclude(['A2'], 'manual check')
    state.set_plot_selection('Target', ['Ctrl-1', 'Ctrl-2', 'Model-1'],
                             ['Ctrl-1', 'Ctrl-2', 'Model-1', 'Model-2'])
    path = tmp_path/'test.qpcrproj'; save_project(path, state)
    restored = load_project(path)
    assert len(restored.wells) == 24
    assert restored.excluded_wells['A2'] == 'manual check'
    assert restored.exclusions[-1].source == 'manual'
    assert restored.reference_genes() == ['REF']
    assert restored.plot_selections['Target'].exploratory is True


def test_column_sample_axis_survives_project_round_trip(tmp_path):
    source = tmp_path / 'source.csv'
    source.write_text('Pos,Cp\nA1,20.0\n', encoding='utf-8')
    wells = tuple(WellRecord(f'{row}{column}', 20.0) for row in ('A', 'B') for column in (1, 2))
    assignments = {}
    for row in ('A', 'B'):
        assignments[f'{row}1'] = PlateAssignment(
            f'{row}1', 'Ctrl-1', 'Ctrl', 'Target', False,
            1 if row == 'A' else 2, plate_row=row)
        assignments[f'{row}2'] = PlateAssignment(
            f'{row}2', 'Model-1', 'Model', 'Target', False,
            1 if row == 'A' else 2, plate_row=row)
    state = ProjectState(
        source_path=str(source), wells=wells, assignments=assignments,
        config=AnalysisConfig(96, sample_axis='column'),
        genes=[GeneDefinition('Target', False, 0)])
    path = tmp_path / 'column.qpcrproj'
    save_project(path, state)
    restored = load_project(path)
    assert restored.config.sample_axis == 'column'
    assert {restored.assignments[f'{row}1'].sample for row in ('A', 'B')} == {'Ctrl-1'}
    assert {restored.assignments[f'{row}2'].sample for row in ('A', 'B')} == {'Model-1'}


def test_exact_well_mapping_survives_project_round_trip_without_renumbering(tmp_path):
    source = tmp_path / 'source.csv'
    source.write_text('Pos,Cp\nB10,20.0\nD10,20.1\n', encoding='utf-8')
    wells = tuple(WellRecord(well, 20.0) for well in ('B10', 'B11', 'B12', 'D10', 'D11', 'D12'))
    assignments = {}
    for row, sample in (('B', 'Low-2'), ('D', 'Low-5')):
        for replicate, column in enumerate(range(10, 13), 1):
            well = f'{row}{column}'
            assignments[well] = PlateAssignment(
                well, sample, 'Low', 'Target', False, replicate, plate_row=row)
    state = ProjectState(
        source_path=str(source), wells=wells, assignments=assignments,
        config=AnalysisConfig(384, sample_axis='row', mapping_scope='well'),
        genes=[GeneDefinition('Target', False, 0)])
    path = tmp_path / 'exact.qpcrproj'
    save_project(path, state)
    restored = load_project(path)
    assert restored.config.mapping_scope == 'well'
    assert {restored.assignments[f'B{column}'].sample for column in range(10, 13)} == {'Low-2'}
    assert {restored.assignments[f'D{column}'].sample for column in range(10, 13)} == {'Low-5'}


def test_custom_group_order_survives_project_round_trip(tmp_path):
    source = tmp_path / 'source.csv'; source.write_text('Pos,Cp\nA1,20.0\n', encoding='utf-8')
    state = ProjectState(
        source_path=str(source), wells=(WellRecord('A1', 20.0),),
        assignments={'A1': PlateAssignment('A1', 'Control-1', 'Control', 'Target')},
        config=AnalysisConfig(
            96, calibrator_group='Control', group_order=['Control', 'Drug']),
        genes=[GeneDefinition('Target', False, 0)])
    path = tmp_path / 'custom-groups.qpcrproj'; save_project(path, state)
    restored = load_project(path)
    assert restored.config.group_order == ['Control', 'Drug']
    assert restored.config.calibrator_group == 'Control'


def test_v1_project_group_migration(tmp_path):
    payload = {'schema_version': 1, 'wells': [{'well':'A1', 'ct':20.0}],
        'assignments': {'A1': {'well':'A1', 'sample':'old', 'group':'Control',
            'gene':'Target', 'is_reference':False, 'technical_replicate':1, 'role':'sample'}},
        'excluded_wells': {'A1':'legacy'}}
    path = tmp_path/'legacy.qpcrproj'
    with zipfile.ZipFile(path, 'w') as archive:
        archive.writestr('project.json', json.dumps(payload))
    state = load_project(path)
    assert state.assignments['A1'].group == 'Ctrl'
    assert state.assignments['A1'].sample == 'Ctrl-1'
    assert state.exclusions[0].source == 'legacy'


def test_clean_excel_formulas_comments_and_svg(tmp_path):
    state = sample_state(tmp_path); state.exclude(['A2'], '', 'qc_suggestion')
    result = analyze(state.wells, state.assignments, state.config, state.excluded_wells)
    excel = tmp_path/'cleaned.xlsx'; export_clean_excel(excel, state)
    workbook = load_workbook(excel, data_only=False)
    assert workbook.sheetnames == ['Target', '排除记录']
    sheet = workbook['Target']
    assert sheet['E2'].value == '=C2-D2'
    assert str(sheet['F2'].value).startswith('=AVERAGE(')
    assert sheet['G2'].value == '=E2-F2'
    assert sheet['H2'].value == '=POWER(2,-G2)'
    assert 'A1=' in sheet['C2'].comment.text and 'A2=' in sheet['C2'].comment.text
    assert workbook['排除记录']['A2'].value == 'A2'
    paths = export_all_svg(tmp_path/'svg', state, result)
    assert len(paths) == 1 and paths[0].suffix == '.svg' and paths[0].stat().st_size > 100


def test_clean_excel_omits_analysis_filtered_samples(tmp_path):
    state = sample_state(tmp_path)
    all_samples = ['Ctrl-1', 'Ctrl-2', 'Model-1', 'Model-2']
    state.set_plot_selection('Target', all_samples[:-1], all_samples)
    excel = tmp_path/'filtered.xlsx'; export_clean_excel(excel, state)
    workbook = load_workbook(excel, data_only=False)
    sheet = workbook['Target']
    exported_samples = [sheet.cell(row, 2).value for row in range(2, sheet.max_row + 1)]
    assert exported_samples == ['Ctrl-1', 'Ctrl-2', 'Model-1']
    assert sheet['F2'].value == '=AVERAGE(E2:E3)'


def test_exploratory_figure_does_not_embed_filter_status(tmp_path):
    state = sample_state(tmp_path)
    chosen = {'Ctrl-1', 'Ctrl-2', 'Model-1'}
    result = analyze(
        state.wells, state.assignments, state.config,
        sample_filters={'Target': chosen},
    )
    figure = make_gene_figure(result, 'Target')
    assert figure._suptitle is None
    all_text = ' '.join(text.get_text() for axis in figure.axes for text in axis.texts)
    assert '已按排除样本重算' not in all_text
    assert 'Ctrl n=' not in all_text
    assert figure.axes[0].get_title() == ''
    assert figure.axes[0].get_ylabel() == 'Relative Target mRNA'
    assert not any(line.get_visible() for line in figure.axes[0].get_ygridlines())
